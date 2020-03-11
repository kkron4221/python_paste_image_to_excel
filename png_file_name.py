from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import glob, os, re, cv2

'''
Description to shorten the interval of pasting 
when the already pasted image is small
'''
def check_img_size(img):
    img_check = cv2.imread(img)
    img_check_hight = img_check.shape[0]
    cell_interval = 0
    if img_check_hight < 500:
        cell_interval = 20
    elif img_check_hight < 600:
        cell_interval = 23
    elif img_check_hight < 700:
        cell_interval = 26
    elif img_check_hight < 800:
        cell_interval = 30
    elif img_check_hight < 900:
        cell_interval = 33
    else:
        print("ERROR 予期しないサイズのファイルがあります。")
    return cell_interval

def paste_image(ws2, right_num_groups):
    before_cell_num = 0
    for num in range(len(right_num_groups)):
        capture_num = 2
        if num != 0:
            cell_size = check_img_size(right_num_groups[num - 1])
            capture_num = before_cell_num + cell_size + 2
        capture_area = 'A' + str(capture_num)
        before_cell_num = capture_num
        ws2.add_image(Image(right_num_groups[num]), capture_area)

def connect_excel(right_num_groups, folder_num):
    wb = load_workbook(filename = excel_name)
    sheet_name = str(folder_num + 1) + "-" + str(sheet_name_list[1]) + "-" + str(sheet_name_list[2])
    sheet_name = sheet_name.rstrip("\'>[0-9]")
    ws2 = wb.create_sheet(title = sheet_name)
    ws2['A1'] = "※画面キャプチャーを添付する"
    paste_image(ws2, right_num_groups)
    wb.save(filename = excel_name)

# TODO: change excel's path and evidence folder path
excel_name = 'test.xlsx'
evidence_folder = "C:\\Users\\username\\Documents\\evidence"
folder_list = os.listdir(evidence_folder)
re_word = r'([0-9]-){2}[0-9]'
check_list = []
for folder_num in range(len(folder_list)):
    num_folder = evidence_folder + "\\" + folder_list[folder_num]
    png_list = os.listdir(num_folder)
    for png_num in range(len(png_list)):
        re_png_name = re.match(re_word, png_list[png_num])
        sheet_name_list = str(re_png_name).split("-")
        png_re_name = num_folder + "\\" + re_png_name.group() + "*"
        right_num_groups = glob.glob(png_re_name)
        if right_num_groups == check_list:
            continue
        connect_excel(right_num_groups, folder_num)
        print(right_num_groups)
        check_list = right_num_groups
print("Finish!!")