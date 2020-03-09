from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

excel_name = 'test.xlsx'
sheet_name = 'fox'
img = Image('sample.png')
'''
TODO:Separate function
    1.Return file_name function
    2.Connection excel function
    3.Paste function
'''


'''
TODO:About png_num
    example 1-1-1 is left_num, center_num, right_num
    1.left_num == folder's num
    2.
'''
png_num = 2
for i in range(0, png_num):
    i = str(i)
    wb = load_workbook(filename = excel_name)
    sheet_name = "sheet_kun" + i
    ws2 = wb.create_sheet(title = sheet_name)
    ws2['A1'] = "※画面キャプチャーを添付する"
    ws2.add_image(img, 'A2')
    wb.save(filename = excel_name)